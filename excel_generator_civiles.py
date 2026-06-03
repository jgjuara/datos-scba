from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pdf_parser_civiles import HEADERS_11, HEADERS_13, clean_int

# Ancho de columnas estándar para la visualización en Excel
COL_WIDTHS_13 = [39.56, 12.67, 10.44, 10.44, 8.0, 10.44, 10.44, 10.44, 11.56, 11.56, 11.56, 3.33, 6.89]
COL_WIDTHS_11 = COL_WIDTHS_13[:11]


def create_excel(data, output_path):
    """
    Crea un libro Excel (.xlsx) con openpyxl aplicando los estilos exactos
    del archivo de muestra: anchos, altos, bordes, alineación y fuentes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Table 1"
    
    num_cols = data["num_cols"]
    headers = HEADERS_13 if num_cols == 13 else HEADERS_11
    widths = COL_WIDTHS_13 if num_cols == 13 else COL_WIDTHS_11
    
    # Configurar anchos de columna
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
        
    # Estilos de borde y fuentes
    thin_side = Side(style='thin', color='FF000000')
    border_grid = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    font_title = Font(name="Arial", size=12.5, bold=True)
    font_subtitle = Font(name="Arial", size=10.5, bold=True)
    font_group = Font(name="Arial", size=8.5, bold=True)
    font_header = Font(name="Arial", size=7.5, bold=True)
    font_data_dept = Font(name="Arial", size=7.0, bold=True)
    font_data_num = Font(name="Arial", size=9.0, bold=False)
    font_total_dept = Font(name="Arial", size=7.0, bold=True)
    font_total_num = Font(name="Arial", size=9.0, bold=True)
    font_avg_dept = Font(name="Arial", size=8.0, bold=True)
    font_avg_num = Font(name="Arial", size=9.0, bold=True)
    font_footnote = Font(name="Arial", size=6.0, bold=True)
    
    # Escribir títulos iniciales
    ws.row_dimensions[1].height = 11.0
    
    ws.row_dimensions[2].height = 17.25
    ws.cell(row=2, column=1, value=data["title1"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws.cell(row=2, column=1).font = font_title
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[3].height = 11.0
    
    ws.row_dimensions[4].height = 15.0
    ws.cell(row=4, column=1, value=data["title2"])
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=11)
    ws.cell(row=4, column=1).font = font_subtitle
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[5].height = 11.0
    
    ws.row_dimensions[6].height = 15.0
    ws.cell(row=6, column=3, value=data["group_header"])
    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=10)
    ws.cell(row=6, column=3).font = font_group
    ws.cell(row=6, column=3).alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir cabecera de columnas
    ws.row_dimensions[7].height = 57.0
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=c_idx, value=header)
        cell.font = font_header
        align = "left" if c_idx == 1 else "center"
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = border_grid
        
    ws.row_dimensions[8].height = 11.0
    for c_idx in range(1, num_cols + 1):
        ws.cell(row=8, column=c_idx).border = border_grid
        
    # Escribir filas de datos de departamentos
    current_row = 9
    for r_data in data["data_rows"]:
        ws.row_dimensions[current_row].height = 12.75
        for c_idx in range(1, num_cols + 1):
            val = r_data[c_idx - 1] if c_idx - 1 < len(r_data) else None
            cell = ws.cell(row=current_row, column=c_idx)
            cell.border = border_grid
            if c_idx == 1:
                cell.value = val
                cell.font = font_data_dept
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.value = clean_int(val)
                cell.font = font_data_num
                cell.alignment = Alignment(horizontal="right", vertical="top")
        current_row += 1
        
    # Fila vacía antes del total
    ws.row_dimensions[current_row].height = 11.0
    for c_idx in range(1, num_cols + 1):
        ws.cell(row=current_row, column=c_idx).border = border_grid
    current_row += 1
    
    # Escribir total provincial
    ws.row_dimensions[current_row].height = 12.75
    for c_idx in range(1, num_cols + 1):
        val = data["total_row"][c_idx - 1] if data["total_row"] and c_idx - 1 < len(data["total_row"]) else None
        cell = ws.cell(row=current_row, column=c_idx)
        cell.border = border_grid
        if c_idx == 1:
            cell.value = val
            cell.font = font_total_dept
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            cell.value = clean_int(val)
            cell.font = font_total_num
            cell.alignment = Alignment(horizontal="right", vertical="top")
    current_row += 1
    
    # Fila vacía antes del promedio
    ws.row_dimensions[current_row].height = 11.0
    for c_idx in range(1, num_cols + 1):
        ws.cell(row=current_row, column=c_idx).border = border_grid
    current_row += 1
    
    # Escribir promedio por juzgado
    ws.row_dimensions[current_row].height = 12.75
    for c_idx in range(1, num_cols + 1):
        val = data["average_row"][c_idx - 1] if data["average_row"] and c_idx - 1 < len(data["average_row"]) else None
        cell = ws.cell(row=current_row, column=c_idx)
        cell.border = border_grid
        if c_idx == 1:
            cell.value = val
            cell.font = font_avg_dept
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            cell.value = clean_int(val)
            cell.font = font_avg_num
            cell.alignment = Alignment(horizontal="right", vertical="top")
    current_row += 1
    
    # Fila vacía antes de las notas al pie
    ws.row_dimensions[current_row].height = 11.0
    for c_idx in range(1, num_cols + 1):
        ws.cell(row=current_row, column=c_idx).border = border_grid
    current_row += 1
    
    # Escribir notas al pie
    for idx, note in enumerate(data["footnotes"]):
        h = 11.25 if idx == 2 else 11.0
        ws.row_dimensions[current_row].height = h
        
        cell_a = ws.cell(row=current_row, column=1, value=note)
        cell_a.font = font_footnote
        cell_a.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        if idx == 0:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        elif idx == 1:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            
        for c_idx in range(1, num_cols + 1):
            ws.cell(row=current_row, column=c_idx).border = border_grid
            
        current_row += 1
        
    wb.save(output_path)
