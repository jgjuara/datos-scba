import os
import re
import csv
import sys
import argparse
import openpyxl

CSV_HEADERS = [
    "Departamento / Sede",
    "Ingresadas",
    "Sentencia",
    "Conciliación",
    "Allanamiento",
    "Transacción",
    "Caducidad",
    "Desistimiento",
    "Interlocutorios",
    "Incompetencia",
    "Total Resueltas",
    "Anio"
]


def infer_year(file_path: str, sheet) -> int:
    """
    Infiere el año del reporte judicial.
    Primero busca un patrón de 4 dígitos en el nombre del archivo.
    Si no lo encuentra, busca un patrón similar en las primeras 10 filas de la hoja.

    Precondiciones:
        - file_path es una cadena válida que representa una ruta de archivo.
        - sheet es una instancia válida de una hoja de openpyxl.

    Postcondiciones:
        - Retorna el año extraído como entero.

    Excepciones:
        - ValueError: Se lanza si no se encuentra ningún año válido de 4 dígitos.
    """
    filename = os.path.basename(file_path)
    match = re.search(r"(20\d{2})", filename)
    if match:
        return int(match.group(1))

    # Inspeccionar las celdas del bloque de título
    for r_idx in range(1, 11):
        for c_idx in range(1, 10):
            val = sheet.cell(row=r_idx, column=c_idx).value
            if val and isinstance(val, str):
                year_match = re.search(r"(20\d{2})", val)
                if year_match:
                    return int(year_match.group(1))

    raise ValueError(f"No se pudo inferir el año del archivo: {file_path}")


def clean_cell_value(val):
    """
    Limpia y normaliza el valor de una celda de Excel.
    Elimina espacios en blanco adicionales y convierte cadenas numéricas a enteros.

    Precondiciones:
        - val puede ser de cualquier tipo (int, str, None, etc.).

    Postcondiciones:
        - Retorna un entero si la cadena contiene solo dígitos (removiendo puntos decimales de miles).
        - Retorna None para valores vacíos o guiones.
        - Retorna el valor en texto limpio en cualquier otro caso.
    """
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str in ["", "-", "None"]:
        return None

    # Eliminar puntos de separación de miles y espacios para verificar si es entero
    normalized_val = val_str.replace(".", "").replace(" ", "")
    if normalized_val.isdigit():
        return int(normalized_val)

    return val_str


def export_xlsx_to_csv(xlsx_path: str, csv_path: str) -> None:
    """
    Carga un archivo Excel de juzgados civiles o tribunales de trabajo,
    detecta la columna inicial de la tabla, extrae las filas de datos
    y las guarda en formato CSV con el año correspondiente.

    Precondiciones:
        - xlsx_path debe ser la ruta de un archivo Excel existente.
        - csv_path debe ser una ruta de destino de escritura válida.

    Postcondiciones:
        - Crea el archivo CSV en la ruta especificada.

    Excepciones:
        - FileNotFoundError: Se lanza si xlsx_path no existe.
        - ValueError: Se lanza si no se encuentra la cabecera 'Departamento' o 'Departamento / Sede'.
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"El archivo Excel de origen no existe: {xlsx_path}")

    # data_only=True asegura que leamos los valores finales evaluados y no las fórmulas
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb.active

    year = infer_year(xlsx_path, sheet)
    rows_to_write = []
    header_found = False
    start_col_idx = 0

    for row in sheet.iter_rows(values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        # Buscar la cabecera dinámica si aún no se ha encontrado
        if not header_found:
            for idx, cell in enumerate(row):
                if cell and "Departamento" in str(cell):
                    header_found = True
                    start_col_idx = idx
                    break
            continue

        # Si ya se encontró la cabecera, procesamos la fila usando start_col_idx
        if len(row) <= start_col_idx:
            continue

        first_cell = str(row[start_col_idx]).strip() if row[start_col_idx] is not None else ""

        # Detener la lectura al alcanzar las notas al pie de página
        if any(foot in first_cell for foot in ["Fuente:", "Los datos son", "No incluye", "(*)"]):
            break

        # Extraer las 11 columnas a partir de la columna inicial detectada
        sliced_row = row[start_col_idx : start_col_idx + 11]
        cleaned_row = [clean_cell_value(cell) for cell in sliced_row]

        # Rellenar con None si la fila es más corta de las 11 columnas requeridas
        while len(cleaned_row) < 11:
            cleaned_row.append(None)

        # Ignorar filas vacías resultantes de la limpieza
        if all(cell is None for cell in cleaned_row):
            continue

        dept_name = cleaned_row[0]
        if dept_name is None:
            continue

        # Agregar el año inferido
        cleaned_row.append(year)
        rows_to_write.append(cleaned_row)

    if not header_found:
        raise ValueError(f"No se encontró la cabecera en el archivo Excel: {xlsx_path}")

    # Crear el directorio del CSV de salida si no existe
    output_dir = os.path.dirname(csv_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    with open(csv_path, mode="w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(CSV_HEADERS)
        writer.writerows(rows_to_write)


def main():
    """
    Punto de entrada principal. Configura los argumentos de línea de comandos
    y maneja el flujo de ejecución y excepciones.
    """
    parser = argparse.ArgumentParser(
        description="Exporta datos de juzgados/tribunales desde un archivo Excel a CSV."
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        default=os.path.join("xlsx", "juzgados_civiles_anual_2025.xlsx"),
        help="Ruta al archivo Excel de origen"
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Ruta al archivo CSV de salida (default: carpeta 'csv' con el mismo nombre de archivo)"
    )

    args = parser.parse_args()

    # Resolver ruta de salida por defecto si no se especificó
    if not args.output:
        base_name = os.path.basename(args.input_file)
        csv_name = base_name.replace(".xlsx", ".csv")
        args.output = os.path.join("csv", csv_name)

    print(f"Iniciando exportación: {args.input_file} -> {args.output}")

    try:
        export_xlsx_to_csv(args.input_file, args.output)
        print("Exportación completada exitosamente.")
    except Exception as e:
        print(f"Error durante la exportación: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
