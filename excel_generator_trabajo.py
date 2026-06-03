from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, DEFAULT_FONT
from openpyxl.utils import get_column_letter

# Ancho de columnas estándar para la visualización en Excel de tribunales de trabajo
COL_WIDTHS_TRABAJO = [
    9.33, 33.78, 10.44, 11.56, 11.56, 8.0, 8.0, 9.33, 8.0, 9.33, 8.0, 10.44, 4.67, 5.78, 22.0
]

def clean_val(val):
    """
    Limpia y convierte un valor a entero o mantiene guiones/cadenas.
    Elimina espacios y puntos de miles.
    Retorna None para valores vacíos o 'None'.
    """
    if val is None:
        return None
    val_str = str(val).strip().replace(" ", "")
    if val_str in ["", "None"]:
        return None
    if val_str == "-":
        return "-"
    try:
        return int(val_str.replace(".", ""))
    except ValueError:
        return val_str

def create_excel_trabajo(data: dict, output_path: str):
    """
    Crea un libro Excel (.xlsx) para Tribunales de Trabajo aplicando los estilos
    exactos del archivo de muestra: anchos, altos, bordes, alineación y fuentes.
    """
    # Establecer la fuente por defecto global del libro de trabajo antes de instanciarlo
    DEFAULT_FONT.name = "Times New Roman"
    DEFAULT_FONT.size = 10.0
    DEFAULT_FONT.bold = False

    wb = Workbook()
    ws = wb.active
    ws.title = "Table 1"

    # Configurar anchos de columna para A a O
    for idx, w in enumerate(COL_WIDTHS_TRABAJO, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    max_row = 9 + len(data["data_rows"]) + 8

    # Estilos de borde y fuentes específicas
    thin_side = Side(style='thin', color='FF000000')
    border_grid = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    font_title = Font(name="Arial", size=12.0, bold=True)
    font_subtitle = Font(name="Arial", size=10.5, bold=True)
    font_group = Font(name="Arial", size=8.5, bold=True)
    font_header = Font(name="Arial", size=7.0, bold=True)
    font_data_dept = Font(name="Arial", size=7.0, bold=True)
    font_data_num = Font(name="Arial", size=8.5, bold=False)
    font_total_dept = Font(name="Arial", size=8.0, bold=True)
    font_total_num = Font(name="Arial", size=8.5, bold=True)
    font_avg_dept = Font(name="Arial", size=8.0, bold=True)
    font_avg_num = Font(name="Arial", size=8.5, bold=True)
    font_footnote_bold = Font(name="Arial", size=6.0, bold=True)
    font_footnote_reg = Font(name="Arial", size=7.0, bold=False)

    # Configurar la altura y alineación vertical de las filas
    for r in range(1, max_row + 1):
        if r == 2:
            row_valign = "center"
        elif r in (7, 45):
            row_valign = "top"
        else:
            row_valign = "bottom"
        ws.row_dimensions[r].alignment = Alignment(horizontal="left", vertical=row_valign, wrap_text=True)

    def apply_row_borders(row_idx):
        for col in range(2, 15):  # Columnas B a N (2 a 14)
            ws.cell(row=row_idx, column=col).border = border_grid

    # Fila 1 (Vacía con bordes)
    ws.row_dimensions[1].height = 10.5
    apply_row_borders(1)

    # Fila 2: Título principal
    ws.row_dimensions[2].height = 17.25
    ws.cell(row=2, column=2, value=data["title1"])
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=12)
    ws.cell(row=2, column=2).font = font_title
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    apply_row_borders(2)

    # Fila 3 (Vacía con bordes)
    ws.row_dimensions[3].height = 11.0
    apply_row_borders(3)

    # Fila 4: Subtítulo
    ws.row_dimensions[4].height = 15.0
    ws.cell(row=4, column=2, value=data["title2"])
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=12)
    ws.cell(row=4, column=2).font = font_subtitle
    ws.cell(row=4, column=2).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    apply_row_borders(4)

    # Fila 5 (Vacía con bordes)
    ws.row_dimensions[5].height = 11.0
    apply_row_borders(5)

    # Fila 6: Encabezado de grupo
    ws.row_dimensions[6].height = 14.0
    ws.cell(row=6, column=4, value=data["group_header"])
    ws.merge_cells(start_row=6, start_column=4, end_row=6, end_column=11)
    ws.cell(row=6, column=4).font = font_group
    ws.cell(row=6, column=4).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    apply_row_borders(6)

    # Fila 7: Cabeceras de columnas
    ws.row_dimensions[7].height = 57.0
    headers = [
        "Departamento / Sede", "Ingresadas", "Sentencia", "Conciliación",
        "Allanamiento", "Transacción", "Caducidad", "Desistimiento",
        "Interlocutorios", "Incompetencia", "Total Resueltas", None, "Cant Trib."
    ]
    for c_idx, header in enumerate(headers, 2):  # Escribir en columnas B a N (2 a 14)
        if c_idx == 13:  # Omitir dar formato personalizado a la columna M (columna espaciadora)
            continue
        cell = ws.cell(row=7, column=c_idx, value=header)
        cell.font = font_header
        if c_idx == 2:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    apply_row_borders(7)

    # Fila 8 (Vacía con bordes)
    ws.row_dimensions[8].height = 11.0
    apply_row_borders(8)

    # Filas de datos (desde la Fila 9)
    current_row = 9
    for r_data in data["data_rows"]:
        ws.row_dimensions[current_row].height = 12.0
        apply_row_borders(current_row)
        for c_idx in range(1, 14):  # Procesar 13 elementos del PDF
            excel_col = c_idx + 1  # Escribir en B a N
            if excel_col == 13:  # Omitir columna espaciadora M
                continue
            val = r_data[c_idx - 1] if c_idx - 1 < len(r_data) else None
            cell = ws.cell(row=current_row, column=excel_col)
            
            if excel_col == 2:
                cell.value = str(val).strip() if val is not None else None
                cell.font = font_data_dept
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cleaned = clean_val(val)
                if cleaned is not None:
                    cell.value = cleaned
                    cell.font = font_data_num
                    if cleaned == "-":
                        align_h = "left" if excel_col in (6, 9) else ("center" if excel_col in (7, 8, 10) else "right")
                        cell.alignment = Alignment(horizontal=align_h, vertical="top", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="top")
        current_row += 1

    # Fila vacía antes del total
    ws.row_dimensions[current_row].height = 11.0
    apply_row_borders(current_row)
    current_row += 1

    # Fila de Total Provincial
    ws.row_dimensions[current_row].height = 12.0
    apply_row_borders(current_row)
    for c_idx in range(1, 14):
        excel_col = c_idx + 1
        if excel_col == 13:  # Omitir columna espaciadora M
            continue
        val = data["total_row"][c_idx - 1] if data["total_row"] and c_idx - 1 < len(data["total_row"]) else None
        cell = ws.cell(row=current_row, column=excel_col)
        
        if excel_col == 2:
            cell.value = str(val).strip() if val is not None else None
            cell.font = font_total_dept
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            cleaned = clean_val(val)
            if cleaned is not None:
                cell.value = cleaned
                cell.font = font_total_num
                cell.alignment = Alignment(horizontal="right", vertical="top")
    current_row += 1

    # Fila vacía antes del promedio
    ws.row_dimensions[current_row].height = 11.0
    apply_row_borders(current_row)
    current_row += 1

    # Fila de Promedio por Tribunal
    ws.row_dimensions[current_row].height = 12.0
    apply_row_borders(current_row)
    for c_idx in range(1, 14):
        excel_col = c_idx + 1
        if excel_col == 13:  # Omitir columna espaciadora M
            continue
        val = data["average_row"][c_idx - 1] if data["average_row"] and c_idx - 1 < len(data["average_row"]) else None
        cell = ws.cell(row=current_row, column=excel_col)
        
        if excel_col == 2:
            cell.value = str(val).strip() if val is not None else None
            cell.font = font_avg_dept
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            cleaned = clean_val(val)
            if cleaned is not None:
                cell.value = cleaned
                cell.font = font_avg_num
                cell.alignment = Alignment(horizontal="right", vertical="top")
    current_row += 1

    # Fila vacía antes de las notas al pie
    ws.row_dimensions[current_row].height = 11.0
    apply_row_borders(current_row)
    current_row += 1

    # Escribir notas al pie (generalmente 3 notas)
    for idx, note in enumerate(data["footnotes"]):
        h = 11.25 if idx == 2 else 11.0
        ws.row_dimensions[current_row].height = h
        apply_row_borders(current_row)
        
        cell_b = ws.cell(row=current_row, column=2, value=note)
        if idx == 1:
            cell_b.font = font_footnote_reg
        else:
            cell_b.font = font_footnote_bold
            
        cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Aplicar combinaciones específicas para las notas
        if idx == 0:
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        elif idx == 1:
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

        current_row += 1

    # Escribir pie de página institucional final en A45 (combinado A45:O45)
    ws.row_dimensions[current_row].height = 16.25
    footer_text = f"Area de Estadísticas - Secretaría de Planificación" + " " * 79 + "Iniciadas y resueltas" + " " * 143 + data["date_str"]
    cell_a = ws.cell(row=current_row, column=1, value=footer_text)
    cell_a.font = Font(name="Arial", size=7.0, bold=False)
    cell_a.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)

    # Pase final de estilos para garantizar fuentes y alineaciones explícitas en el XML
    default_font = Font(name="Times New Roman", size=10.0, bold=False)
    for r in range(1, max_row + 1):
        if r == 2:
            row_valign = "center"
        elif r in (7, 45):
            row_valign = "top"
        else:
            row_valign = "bottom"
            
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            # Reemplazar fuente por defecto de Calibri
            if not cell.font or cell.font.name != "Arial":
                cell.font = default_font
            # Rellenar alineación explícita si está vacía o parcialmente vacía
            if not cell.alignment or (cell.alignment.horizontal is None and cell.alignment.vertical is None):
                cell.alignment = Alignment(horizontal="left", vertical=row_valign, wrap_text=True)
            elif cell.alignment:
                horiz = cell.alignment.horizontal or "left"
                vert = cell.alignment.vertical or row_valign
                wrap = cell.alignment.wrap_text if cell.alignment.wrap_text is not None else True
                cell.alignment = Alignment(horizontal=horiz, vertical=vert, wrap_text=wrap)

    wb.save(output_path)
